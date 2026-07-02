#!/usr/bin/env python3
import os
import sys
import argparse
import logging
from dotenv import load_dotenv
import web_driver_utils
from web_agent import WebAgent
from web_reporter import WebHTMLReporter

# Load environment configuration (.env)
load_dotenv()

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("web_agent.log", mode="w")
    ]
)
logger = logging.getLogger("WebAgent.main")

def sanitize_class_name(s: str) -> str:
    """Sanitize description strings into valid PascalCase Java class identifiers."""
    import re
    # Remove special characters, split by spaces/dashes/underscores
    words = re.split(r'[^a-zA-Z0-9]+', s)
    capitalized_words = [w.capitalize() for w in words if w]
    name = "".join(capitalized_words)
    if not name:
        return "PlaywrightAutomationCase"
    # Ensure it starts with a letter (prepend 'TC' if it starts with a number)
    if name[0].isdigit():
        name = "TC" + name
    return name

def parse_excel_test_cases(file_path: str) -> list:
    """Parse multiple rows of step-by-step instructions from a standard manual Excel sheet."""
    import openpyxl
    import re
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    step_headers = {"test steps", "test step", "steps", "step description", "actions", "action", "step"}
    desc_headers = {"description", "test case description", "name", "test case name", "title", "summary", "test case"}
    
    step_col_idx = None
    desc_col_idx = None
    header_row_idx = None
    
    # Inspect first 10 rows to find headers
    for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), 1):
        for col_idx, val in enumerate(row, 1):
            if val:
                val_clean = str(val).strip().lower()
                if val_clean in step_headers and step_col_idx is None:
                    step_col_idx = col_idx
                    header_row_idx = row_idx
                elif val_clean in desc_headers and desc_col_idx is None:
                    desc_col_idx = col_idx
                    header_row_idx = row_idx
        if step_col_idx is not None:
            break
            
    test_cases = []
    
    # Helper to clean individual steps
    def clean_step(s: str) -> str:
        s_clean = s.strip()
        s_clean = re.sub(r'^(?i:step\s+\d+[\s\.:\-]*)?[\d\-\*\•\.\)\(]*\s*', '', s_clean).strip()
        return s_clean
        
    def get_steps_from_value(val) -> list:
        if val is None:
            return []
        val_str = str(val).strip()
        if not val_str or val_str.startswith("#"):
            return []
            
        steps = []
        lines = val_str.splitlines()
        for line in lines:
            line_clean = clean_step(line)
            if line_clean and not line_clean.startswith("#"):
                steps.append(line_clean)
        return steps

    # If steps column was identified
    if step_col_idx is not None:
        start_row = header_row_idx + 1
        for r in range(start_row, sheet.max_row + 1):
            steps_val = sheet.cell(row=r, column=step_col_idx).value
            row_steps = get_steps_from_value(steps_val)
            if not row_steps:
                continue
                
            desc_val = None
            if desc_col_idx is not None:
                desc_val = sheet.cell(row=r, column=desc_col_idx).value
                
            if desc_val:
                desc_str = str(desc_val).strip()
            else:
                desc_str = f"TestCase_{r - start_row + 1}"
                
            test_cases.append({
                "name": sanitize_class_name(desc_str),
                "description": desc_str,
                "steps": row_steps,
                "source": f"{os.path.basename(file_path)}: Row {r}"
            })
    else:
        # Fallback: scan first 5 columns to find the first one that has text values in first 10 rows
        fallback_col_idx = 1
        for col_idx in range(1, min(6, sheet.max_column + 1) if sheet.max_column else 6):
            has_text = False
            for row in sheet.iter_rows(max_row=10, min_col=col_idx, max_col=col_idx, values_only=True):
                val = row[0]
                if val and isinstance(val, str) and len(val.strip()) > 3:
                    has_text = True
                    break
            if has_text:
                fallback_col_idx = col_idx
                break
                
        for r in range(1, sheet.max_row + 1):
            val = sheet.cell(row=r, column=fallback_col_idx).value
            if val is not None:
                val_str = str(val).strip()
                if val_str.lower() in step_headers:
                    continue
                row_steps = get_steps_from_value(val)
                if row_steps:
                    test_cases.append({
                        "name": f"TestCase_{r}",
                        "description": f"TestCase_{r}",
                        "steps": row_steps,
                        "source": f"{os.path.basename(file_path)}: Row {r}"
                    })
                    
    wb.close()
    return test_cases

def load_all_test_cases(tc_paths: list) -> list:
    """Load all test cases from text files or Excel spreadsheets."""
    test_cases = []
    for path in tc_paths:
        abs_path = os.path.abspath(path)
        if not os.path.exists(abs_path):
            logger.error(f"Error: Testcase file does not exist at '{abs_path}'")
            sys.exit(1)
            
        basename = os.path.basename(abs_path)
        name_no_ext, ext = os.path.splitext(basename)
        
        if ext.lower() in [".xlsx", ".xls"]:
            sheet_cases = parse_excel_test_cases(abs_path)
            test_cases.extend(sheet_cases)
        else:
            steps = []
            with open(abs_path, "r", encoding="utf-8") as f:
                for line in f:
                    line_clean = line.strip()
                    if not line_clean or line_clean.startswith("#"):
                        continue
                    clean_step = line_clean.lstrip("-*").strip()
                    if clean_step:
                        steps.append(clean_step)
            if steps:
                test_cases.append({
                    "name": sanitize_class_name(name_no_ext),
                    "description": name_no_ext,
                    "steps": steps,
                    "source": basename
                })
    return test_cases

def main():
    parser = argparse.ArgumentParser(description="WebAgent - Autonomous Web Testcase Playwright Agent")
    parser.add_argument("--url", required=True, help="Initial target URL to navigate to")
    parser.add_argument("--testcase", required=True, help="Path to text or Excel testcase files (comma-separated list)")
    parser.add_argument("--output-java", default="PlaywrightAutomation.java", help="Path to write the generated Playwright Java class file")
    parser.add_argument("--output-report", default="web_report.html", help="Path to write the visual HTML report")
    parser.add_argument("--headless", action="store_true", help="Run browser in headless mode (default: False)")
    parser.add_argument("--model", default="gpt-4o", help="OpenAI Model name (default: gpt-4o)")
    
    args = parser.parse_args()
 
    # Split input comma-separated files
    tc_files = [x.strip() for x in args.testcase.split(",") if x.strip()]
    
    # Load all test cases
    test_cases = load_all_test_cases(tc_files)
    
    if not test_cases:
        logger.error("Error: No valid test cases found.")
        sys.exit(1)
        
    logger.info(f"Loaded {len(test_cases)} test case(s) for execution.")
    
    overall_success = True
    results = []
    test_cases_runs = []
    
    # Initialize the WebAgent
    agent = WebAgent(model=args.model)
    
    for tc_idx, tc in enumerate(test_cases, 1):
        logger.info(f"\n==================================================")
        logger.info(f"RUNNING TEST CASE {tc_idx} of {len(test_cases)}: {tc['name']}")
        logger.info(f"Source: {tc['source']}")
        logger.info(f"Steps ({len(tc['steps'])}):")
        for idx, s in enumerate(tc["steps"], 1):
            logger.info(f"  {idx}. {s}")
        logger.info(f"==================================================")
        
        playwright_instance = None
        browser = None
        page = None
        
        try:
            # Initialize browser & context
            playwright_instance, browser, page = web_driver_utils.setup_browser(headless=args.headless)
            
            # Navigate to initial target URL
            logger.info(f"Navigating to initial target URL: {args.url}")
            page.goto(args.url)
            
            # Execute test case
            success, logs = agent.execute_testcase(page, tc["steps"])
            
            test_cases_runs.append({
                "name": tc["name"],
                "description": tc.get("description", tc["name"]),
                "logs": logs
            })
            
            # Generate Playwright Java code (for the HTML report)
            java_code = agent.generate_java_code(logs, tc["name"], initial_url=args.url)
            
            report_dir = os.path.dirname(os.path.abspath(args.output_report))
            report_filepath = os.path.join(report_dir, f"web_report_{tc['name']}.html")
            
            # Generate HTML report
            reporter = WebHTMLReporter(target_url=args.url, testcase_name=tc["name"])
            report_file = reporter.generate_report(logs, java_code, report_filepath)
            logger.info(f"✓ Automated visual HTML report generated at: {report_file}")
            
            if success:
                logger.info(f"🎉 Test Case '{tc['name']}' executed successfully!")
                results.append((tc["name"], "PASSED", None))
            else:
                logger.warning(f"⚠️ Test Case '{tc['name']}' encountered failures.")
                results.append((tc["name"], "FAILED", "Step failures encountered"))
                overall_success = False
                
        except Exception as e:
            logger.error(f"Error running test case '{tc['name']}': {e}", exc_info=True)
            results.append((tc["name"], "ERROR", str(e)))
            overall_success = False
            
        finally:
            # Clean up browser session
            logger.info("Cleaning up browser sessions...")
            if browser:
                try:
                    browser.close()
                except Exception as e:
                    logger.error(f"Error closing browser: {e}")
            if playwright_instance:
                try:
                    playwright_instance.stop()
                except Exception as e:
                    logger.error(f"Error stopping Playwright: {e}")
            logger.info("Playwright session terminated.")
            
    # Generate unified TestNG Java code at the end
    if test_cases_runs:
        java_filename = os.path.basename(args.output_java)
        class_name, _ = os.path.splitext(java_filename)
        import re
        class_name = re.sub(r'[^a-zA-Z0-9_]', '', class_name)
        if class_name and class_name[0].isdigit():
            class_name = "TC" + class_name
        if not class_name:
            class_name = "PlaywrightAutomation"
        
        testng_code = agent.generate_testng_java_code(test_cases_runs, class_name=class_name, initial_url=args.url)
        
        output_dir = os.path.dirname(os.path.abspath(args.output_java))
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
            
        with open(args.output_java, "w", encoding="utf-8") as f:
            f.write(testng_code)
        logger.info(f"✓ Generated TestNG Playwright Java code written to: {args.output_java}")
        
    # Log overall execution summary
    logger.info(f"\n==================================================")
    logger.info(f"EXECUTION SUMMARY")
    logger.info(f"==================================================")
    for name, status, err in results:
        status_str = f"[{status}]"
        err_str = f" - Error: {err}" if err else ""
        logger.info(f"  * {name:<35} {status_str:<10}{err_str}")
    logger.info(f"==================================================")
    
    if overall_success:
        logger.info("🎉 All test cases completed successfully!")
    else:
        logger.warning("⚠️ Some test cases encountered failures or errors.")
        sys.exit(1)

if __name__ == "__main__":
    main()
